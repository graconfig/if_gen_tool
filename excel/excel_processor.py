"""
SAP IF Process
"""

import shutil
import warnings
import threading
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import pandas as pd

from utils.i18n import _
from utils.sap_logger import logger
from hana.hana_conn import HANADBClient
from models.data_models import InterfaceField
from tqdm import tqdm
from odata.odata import odata_verify

# Suppress the specific DrawingML warning by matching the message text
warnings.filterwarnings(
    "ignore",
    message="DrawingML support is incomplete and limited to charts and images only. Shapes and drawings will be lost.",
    category=UserWarning,
)


class ExcelProcessor:
    def __init__(self, data_dir: Path, ai_service, config_manager):
        self.data_dir = data_dir
        self.ai_service = ai_service
        self.config_manager = config_manager

        self.excel_config = config_manager.get_excel_config()
        self.column_mappings = None

        # Use environment-based configuration
        self.batch_size = int(self.excel_config.get("batch_size", 30))
        self.max_concurrent_batches = int(
            self.excel_config.get("max_concurrent_batches", 5)
        )
        self.config_source = "environment variables"

    def process_file(self, file_path: Path) -> None:
        if not file_path.exists():
            raise FileNotFoundError(_("❌ Input file not found: {}").format(file_path))

        from core.consts import Directories

        current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_stem = file_path.stem
        file_suffix = file_path.suffix

        # 输出文件名
        output_filename = f"processed_{current_date}_{file_stem}{file_suffix}"
        output_path = self.data_dir / Directories.EXCEL_OUTPUT / output_filename

        if output_path.exists():
            output_path.unlink()

        workbook = openpyxl.load_workbook(file_path)

        sheet_name = self.excel_config["sheet_name"]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(_("Sheet '{}' not found in workbook").format(sheet_name))

        self._process_worksheet(workbook, sheet_name, file_path.name)

        workbook.save(output_path)
        logger.info(
            _("Processed file saved: {} ✅").format(output_filename),
            logger.get_excel_log_filename(file_path.name),
        )

        # Archive the source file after successful processing
        if self._archive_processed_file(file_path):
            logger.info(
                _("Source file moved to archive folder."),
                logger.get_excel_log_filename(file_path.name),
            )
        else:
            logger.warning(
                _(
                    "⚠️ Warning: Could not archive source file, but processing completed successfully."
                ),
                logger.get_excel_log_filename(file_path.name),
            )

    def _process_worksheet(
        self, workbook: openpyxl.Workbook, sheet_name: str, excel_filename: str
    ) -> None:
        worksheet = workbook[sheet_name]

        # 提前输入列的字段
        input_fields = self.extract_fields(worksheet)

        # If we have a small number of fields, process normally
        if len(input_fields) <= self.batch_size:
            self._process_single(worksheet, input_fields, excel_filename)
        else:
            # For large files, process in batches
            self._process_in_batches(worksheet, input_fields, excel_filename)

    def _process_single(
        self, worksheet, input_fields: List[InterfaceField], excel_filename: str
    ) -> None:
        """Process a single batch of fields with custom field priority"""
        
        # ========== 新增：客户化字段优先匹配 ==========
        logger.info(
            _("Step 1: Matching custom fields..."),
            logger.get_excel_log_filename(excel_filename),
        )
        
        matched_results, unmatched_fields = self._match_custom_fields(
            input_fields, excel_filename
        )
        
        logger.info(
            _("Custom field matching: {} matched, {} unmatched").format(
                len(matched_results), len(unmatched_fields)
            ),
            logger.get_excel_log_filename(excel_filename),
        )
        
        # ========== 如果全部匹配成功，直接写入结果 ==========
        if not unmatched_fields:
            logger.info(
                _("All fields matched from custom table, skipping CDS view process."),
                logger.get_excel_log_filename(excel_filename),
            )
            self.write_results(worksheet, matched_results)
            return
        
        # ========== 如果有未匹配字段，继续原有流程 ==========
        logger.info(
            _("Step 2: Processing {} unmatched fields with CDS views...").format(
                len(unmatched_fields)
            ),
            logger.get_excel_log_filename(excel_filename),
        )
        
        # 1、根据输入内容查找视图（只处理未匹配字段）
        final_context_for_llm = []
        module = unmatched_fields[0].module
        if_name = unmatched_fields[0].if_name
        if_desc = unmatched_fields[0].if_desc
        module_query = ",".join([module, if_name, if_desc])

        with HANADBClient() as hana_client:
            log_filename = logger.get_excel_log_filename(excel_filename)

            cat_find_by_module = hana_client.run_vector_search(
                query=module_query, k=3, log_filename=log_filename
            )
            if cat_find_by_module.empty:
                logger.warning(
                    _("No categories for module."),
                    logger.get_excel_log_filename(excel_filename),
                )
                self.write_results(worksheet, matched_results)
                return

            category_string = cat_find_by_module.iloc[0]["VIEWCATEGORY"]
            views_find_by_cat = hana_client.get_views(
                category=category_string, log_filename=log_filename
            )
            if views_find_by_cat.empty:
                logger.warning(
                    _("No views for category."),
                    logger.get_excel_log_filename(excel_filename),
                )
                self.write_results(worksheet, matched_results)
                return
            logger.info(
                "-" * 80,
                logger.get_excel_log_filename(excel_filename),
            )
            logger.info(
                _("Found {} candidate CDS views").format(len(views_find_by_cat)),
                logger.get_excel_log_filename(excel_filename),
            )
            # 根据输入内容查找最相关的视图（传入未匹配字段）
            llm_return_views = self._select_relevant_views(
                views_find_by_cat, unmatched_fields, excel_filename
            )
            if not llm_return_views:
                logger.warning(
                    _("No views found by LLM."),
                    logger.get_excel_log_filename(excel_filename),
                )

            logger.info(
                _("LLM selected {} CDS views.").format(len(llm_return_views)),
                logger.get_excel_log_filename(excel_filename),
            )

            # If no views were selected, write matched results and return
            if not llm_return_views:
                logger.error(
                    _("No CDS views selected for unmatched fields."),
                    logger.get_excel_log_filename(excel_filename),
                )
                self.write_results(worksheet, matched_results)
                return
            # 获取视图的字段
            llm_return_views_fields = hana_client.get_fields(
                cds_views=llm_return_views, log_filename=log_filename
            )

            # Get Custom views fields
            custom_views_fields = hana_client.get_custom_fields(
                log_filename=log_filename
            )

            # Merge fields (both are dictionaries)
            for view_name, fields_list in custom_views_fields.items():
                if view_name in llm_return_views_fields:
                    llm_return_views_fields[view_name].extend(fields_list)
                else:
                    llm_return_views_fields[view_name] = fields_list

            # 3. Prepare the final context for the LLM using filtered fields
            llm_return_views_df = views_find_by_cat[
                views_find_by_cat["VIEWNAME"].isin(llm_return_views)
            ]

            # 获取custom视图对应的DataFrame
            custom_views_df = views_find_by_cat[
                views_find_by_cat["VIEWNAME"].isin(custom_views_fields.keys())
            ]

            llm_return_views_df = pd.concat(
                [llm_return_views_df, custom_views_df], ignore_index=True
            )

            final_context_for_llm = self._prepare_llm_context(
                llm_return_views_fields, llm_return_views_df
            )

            logger.info(
                "-" * 80,
                logger.get_excel_log_filename(excel_filename),
            )

            # 根据查找的视图和字段调用LLM匹配（只匹配未匹配字段）
            logger.info(
                _("Calling LLM to match unmatched fields..."),
                logger.get_excel_log_filename(excel_filename),
            )
            try:
                unmatched_batch_results = self._match_fields(
                    unmatched_fields, final_context_for_llm, excel_filename
                )
            except Exception as e:
                logger.error(
                    f"❌ LLM function call failed: {e}",
                    logger.get_excel_log_filename(excel_filename),
                )
                self.write_results(worksheet, matched_results)
                raise RuntimeError(
                    _("Failed to process file due to LLM error: {}").format(e)
                ) from e
            logger.info(
                "-" * 80,
                logger.get_excel_log_filename(excel_filename),
            )

            # ========== 合并结果：已匹配 + 新匹配 ==========
            unmatched_results = list(zip(unmatched_fields, unmatched_batch_results))
            final_results = matched_results + unmatched_results
            
            # 按行号排序，保持原始顺序
            final_results.sort(key=lambda x: x[0].row_index)

            # Write results
            logger.info(
                _("Writing results: {} from custom, {} from CDS").format(
                    len(matched_results), len(unmatched_results)
                ),
                logger.get_excel_log_filename(excel_filename),
            )
            self.write_results(worksheet, final_results)

    def _process_in_batches(
        self, worksheet, input_fields: List[InterfaceField], excel_filename: str
    ) -> None:
        """Process large number of fields in batches with custom field priority"""
        
        logger.info(
            _("Processing {} rows in batches of {}").format(
                len(input_fields), self.batch_size
            ),
            logger.get_excel_log_filename(excel_filename),
        )

        # ========== 新增：客户化字段优先匹配 ==========
        logger.info(
            _("Step 1: Matching custom fields..."),
            logger.get_excel_log_filename(excel_filename),
        )
        
        matched_results, unmatched_fields = self._match_custom_fields(
            input_fields, excel_filename
        )
        
        logger.info(
            _("Custom field matching: {} matched, {} unmatched").format(
                len(matched_results), len(unmatched_fields)
            ),
            logger.get_excel_log_filename(excel_filename),
        )
        
        # ========== 如果全部匹配成功，直接写入结果 ==========
        if not unmatched_fields:
            logger.info(
                _("All fields matched from custom table, skipping CDS view process."),
                logger.get_excel_log_filename(excel_filename),
            )
            self.write_results(worksheet, matched_results)
            return
        
        # ========== 如果有未匹配字段，继续原有批处理流程 ==========
        logger.info(
            _("Step 2: Processing {} unmatched fields with CDS views in batches...").format(
                len(unmatched_fields)
            ),
            logger.get_excel_log_filename(excel_filename),
        )

        # 1. Get common context for all batches (same module/interface)
        module = unmatched_fields[0].module
        if_name = unmatched_fields[0].if_name
        if_desc = unmatched_fields[0].if_desc
        module_query = ",".join([module, if_name, if_desc])

        with HANADBClient() as hana_client:
            log_filename = logger.get_excel_log_filename(excel_filename)

            # Get CDS views once for all batches
            cat_find_by_module = hana_client.run_vector_search(
                query=module_query, k=3, log_filename=log_filename
            )
            if cat_find_by_module.empty:
                logger.warning(
                    _("No categories for module."),
                    logger.get_excel_log_filename(excel_filename),
                )
                self.write_results(worksheet, matched_results)
                return

            category_string = cat_find_by_module.iloc[0]["VIEWCATEGORY"]
            views_find_by_cat = hana_client.get_views(
                category=category_string, log_filename=log_filename
            )
            if views_find_by_cat.empty:
                logger.warning(
                    _("No views for category."),
                    logger.get_excel_log_filename(excel_filename),
                )
                self.write_results(worksheet, matched_results)
                return

            logger.info(
                _("Found {} candidate CDS views").format(len(views_find_by_cat)),
                logger.get_excel_log_filename(excel_filename),
            )

            # Select relevant views once for all batches (传入未匹配字段)
            llm_return_views = self._select_relevant_views(
                views_find_by_cat, unmatched_fields, excel_filename
            )
            if not llm_return_views:
                logger.warning(
                    _("No views found by LLM."),
                    logger.get_excel_log_filename(excel_filename),
                )
                self.write_results(worksheet, matched_results)
                return

            logger.info(
                _("LLM selected {} CDS views.").format(len(llm_return_views)),
                logger.get_excel_log_filename(excel_filename),
            )

            # Get fields from selected views once for all batches
            llm_return_views_fields = hana_client.get_fields(
                cds_views=llm_return_views, log_filename=log_filename
            )

            # Get Custom views fields
            # custom_views_fields = hana_client.get_custom_fields(log_filename=log_filename)

            # Merge fields
            # llm_return_views_fields.extend(custom_views_fields)

            llm_return_views_df = views_find_by_cat[
                views_find_by_cat["VIEWNAME"].isin(llm_return_views)
            ]

            # 获取custom视图对应的DataFrame
            # custom_views_df = views_find_by_cat[
            #     views_find_by_cat["VIEWNAME"].isin(custom_views_fields)
            # ]
            #
            # llm_return_views_df = pd.concat([llm_return_views_df, custom_views_df],ignore_index=True)

            final_context_for_llm = self._prepare_llm_context(
                llm_return_views_fields, llm_return_views_df
            )

            logger.info(
                "-" * 80,
                logger.get_excel_log_filename(excel_filename),
            )

            # 2. Process unmatched fields in batches (with parallel processing)
            unmatched_results = []

            # Split unmatched fields into batches
            batches = []
            for i in range(0, len(unmatched_fields), self.batch_size):
                batch_fields = unmatched_fields[i : i + self.batch_size]
                batches.append((i, batch_fields))

            logger.info(
                _("Processing {} batches for unmatched fields...").format(len(batches)),
                logger.get_excel_log_filename(excel_filename),
            )

            # Process batches in parallel
            with tqdm(
                total=len(batches),
                desc=_("Processing batches"),
                unit="batch",
                ncols=100,
                leave=False,
            ) as pbar:
                with ThreadPoolExecutor(
                    max_workers=self.max_concurrent_batches
                ) as executor:
                    # Submit all batches for processing
                    future_to_batch = {
                        executor.submit(
                            self._process_batch,
                            batch_fields,
                            final_context_for_llm,
                            excel_filename,
                            i,
                        ): (i, len(batch_fields))
                        for i, batch_fields in batches
                    }

                    # Collect results as they complete
                    completed_batches = 0
                    for future in as_completed(future_to_batch):
                        batch_index, batch_size = future_to_batch[future]
                        batch_start = batch_index + 1
                        batch_end = batch_index + batch_size

                        try:
                            batch_results = future.result()
                            # Combine unmatched fields with results for this batch
                            batch_fields = unmatched_fields[
                                batch_index : batch_index + batch_size
                            ]
                            batch_final_results = list(zip(batch_fields, batch_results))
                            unmatched_results.extend(batch_final_results)

                            logger.info(
                                _("Row {}-{} processed successfully").format(
                                    batch_start, batch_end
                                ),
                                logger.get_excel_log_filename(excel_filename),
                            )
                        except Exception as e:
                            logger.error(
                                _("Failed to process row {}-{}: {}").format(
                                    batch_start, batch_end, e
                                ),
                                logger.get_excel_log_filename(excel_filename),
                            )
                            # Continue with other batches even if one fails

                        # Update progress
                        completed_batches += 1
                        pbar.set_postfix_str(
                            _("Completed: {}/{}").format(
                                completed_batches, len(batches)
                            )
                        )
                        pbar.update(1)
                        # Add newline to separate progress bar from logger output
                        print()  # Add newline after progress update

            # ========== 合并结果：已匹配 + 新匹配 ==========
            all_results = matched_results + unmatched_results
            all_results.sort(key=lambda x: x[0].row_index)

            # 3. Write all results
            logger.info(
                _("Writing {} results: {} from custom, {} from CDS").format(
                    len(all_results), len(matched_results), len(unmatched_results)
                ),
                logger.get_excel_log_filename(excel_filename),
            )
            self.write_results(worksheet, all_results)

    def _process_batch(
        self,
        batch_fields: List[InterfaceField],
        context: List[Dict[str, Any]],
        excel_filename: str,
        batch_index: int,
    ) -> List[Dict[str, Any]]:
        """Process a single batch of fields"""
        try:
            # Ensure the current file is set for this worker thread for proper token tracking
            from utils.token_statistics import set_current_file

            set_current_file(excel_filename)

            return self._match_fields(batch_fields, context, excel_filename)
        except Exception as e:
            logger.error(
                _("Error in batch {}: {}").format(batch_index, e),
                logger.get_excel_log_filename(excel_filename),
            )
            # Return empty results for failed batch
            return [{} for _ in batch_fields]

    def _match_custom_fields(
        self,
        input_fields: List[InterfaceField],
        excel_filename: str
    ) -> Tuple[List[Tuple[InterfaceField, Dict]], List[InterfaceField]]:
        """
        优先匹配客户化字段表（基于 SOURCEDESC 向量检索）
        
        Args:
            input_fields: 输入字段列表
            excel_filename: Excel 文件名
            
        Returns:
            (已匹配的字段结果列表, 未匹配的字段列表)
        """
        matched_results = []
        unmatched_fields = []
        
        # 从配置读取阈值
        threshold = self.excel_config.get("custom_field_threshold", 0.75)
        
        with HANADBClient() as hana_client:
            log_filename = logger.get_excel_log_filename(excel_filename)
            
            for field in input_fields:
                # 构建查询文本：字段名 + 字段描述 + 示例值
                query_parts = [
                    field.field_name,
                    field.field_text,
                    field.sample_value
                ]
                query_text = " ".join([str(p).strip() for p in query_parts if p])
                
                # 向量检索客户化字段表
                custom_match = hana_client.search_custom_field_by_vector(
                    field_query=query_text,
                    threshold=threshold,
                    log_filename=log_filename
                )
                
                if custom_match:
                    # 匹配成功，构建结果
                    match_result = {
                        "table_id": custom_match.get("table_name", ""),
                        "field_id": custom_match.get("field_name", ""),
                        "field_name": custom_match.get("field_desc", ""),
                        "key_flag": "○" if custom_match.get("is_key") else "",
                        "obligatory": custom_match.get("obligatory", ""),
                        "data_type": custom_match.get("data_type", ""),
                        "length_total": custom_match.get("length_total", ""),
                        "length_dec": custom_match.get("length_dec", ""),
                        "field_desc": custom_match.get("field_desc", ""),
                        "sample_value": field.sample_value,
                        "match": "Custom",
                        "notes": f"{int(custom_match.get('similarity', 0) * 100)}% - Custom field match",
                    }
                    matched_results.append((field, match_result))
                    
                    logger.debug(
                        f"Row {field.row_index}: Custom match (similarity: {custom_match.get('similarity', 0):.2%})",
                        log_filename
                    )
                else:
                    # 未匹配，加入未匹配列表
                    unmatched_fields.append(field)
        
        return matched_results, unmatched_fields

    def extract_fields(self, worksheet) -> List[InterfaceField]:
        input_fields = []

        # Detect SAP format by checking the detection cell
        input_system_col = self.excel_config.get("input_system_col", "F")
        input_system_row = self.excel_config.get("input_system_row", 6)
        cell_value = worksheet[f"{input_system_col}{input_system_row}"].value

        # Determine which column mappings to use based on cell value
        if cell_value and "SAP" in str(cell_value).upper():
            self.column_mappings = self.config_manager.get_column_mappings_sap()
        else:
            self.column_mappings = self.config_manager.get_column_mappings()

        header_row = self.excel_config["header_row"]
        input_header_cols = self.column_mappings["input_header_cols"]

        # 抬头module、接口信息
        module = worksheet[f"{input_header_cols['module']}{header_row}"].value or ""
        if_name = worksheet[f"{input_header_cols['if_name']}{header_row}"].value or ""
        if_desc = worksheet[f"{input_header_cols['if_desc']}{header_row}"].value or ""

        start_row = self.excel_config["start_row"]
        input_row_cols = self.column_mappings["input_row_cols"]

        for row in range(start_row, (worksheet.max_row or 1000) + 1):
            field_name = worksheet[f"{input_row_cols['field_name']}{row}"].value
            if field_name == '' or field_name == 'e':
                continue

            interface_field = InterfaceField(
                module=module,
                if_name=if_name,
                if_desc=if_desc,
                field_name=str(field_name).strip(),
                key_flag=worksheet[f"{input_row_cols['key_flag']}{row}"].value or "",
                obligatory=worksheet[f"{input_row_cols['obligatory']}{row}"].value
                or "",
                data_type=worksheet[f"{input_row_cols['data_type']}{row}"].value or "",
                field_id=worksheet[f"{input_row_cols['field_id']}{row}"].value or "",
                length_total=worksheet[f"{input_row_cols['length_total']}{row}"].value
                or "",
                length_dec=worksheet[f"{input_row_cols['length_dec']}{row}"].value
                or "",
                field_text=worksheet[f"{input_row_cols['field_text']}{row}"].value
                or "",
                sample_value=worksheet[f"{input_row_cols['sample_value']}{row}"].value
                or "",
                remark=worksheet[f"{input_row_cols['remark']}{row}"].value
                or "",
                verify=worksheet[f"{input_row_cols['verify']}{row}"].value
                or "",
                row_index=row,
            )

            input_fields.append(interface_field)

        return input_fields

    def _select_relevant_views(
        self,
        candidate_views_df: pd.DataFrame,
        input_fields: List[InterfaceField],
        excel_filename: str,
    ) -> List[str]:
        views_prompt = self.ai_service.get_view_selection_prompt(
            candidate_views_df, input_fields
        )

        views_function_schema = self.ai_service.get_view_selection_schema()

        try:
            views_response = self.ai_service.call_with_function(
                views_prompt, views_function_schema
            )

            if views_response:
                return views_response.get("relevant_view_names", [])
            else:
                return []
        except Exception as e:
            logger.debug(
                f"select relevant views failed:{e}",
                logger.get_excel_log_filename(excel_filename),
            )
            return []

    def _prepare_llm_context(
        self,
        all_fields_dict: Dict[str, List[Dict[str, Any]]],
        relevant_views_df: pd.DataFrame,
    ) -> List[Dict[str, Any]]:
        context_list = []
        view_desc_map = pd.Series(
            relevant_views_df.VIEWDESC.values, index=relevant_views_df.VIEWNAME
        ).to_dict()

        for view_name, fields in all_fields_dict.items():
            view_desc = view_desc_map.get(view_name, "")  # Safely get description
            for field_detail in fields:
                context_list.append(
                    {
                        "view_name": view_name,
                        "view_desc": view_desc,
                        "field_name": field_detail.get("field_name", ""),
                        "field_desc": field_detail.get("field_desc", ""),
                        "is_key": field_detail.get("is_key", False),
                        "data_type": field_detail.get("data_type", ""),
                        "length_total": str(field_detail.get("length_total", "")),
                        "length_dec": str(field_detail.get("length_dec", "")),
                    }
                )
        return context_list

    def write_results(
        self, worksheet, results: List[Tuple[InterfaceField, Dict[str, Any]]]
    ) -> None:
        output_columns = self.column_mappings["output_columns"]
        
        processed_count = 0
        for interface_field, match_result in results:
            row = interface_field.row_index
            isverify = interface_field.verify

            if isverify == "" or isverify == "-":
                try:
                    worksheet[f"{output_columns['field_name']}{row}"] = match_result.get(
                        "field_name", ""
                    )  # Field description
                    worksheet[f"{output_columns['field_id']}{row}"] = match_result.get(
                        "field_id", ""
                    )  # Technical field name
                    worksheet[f"{output_columns['key_flag']}{row}"] = match_result.get(
                        "key_flag", ""
                    )
                    worksheet[f"{output_columns['obligatory']}{row}"] = match_result.get(
                        "obligatory", ""
                    )
                    worksheet[f"{output_columns['table_id']}{row}"] = match_result.get(
                        "table_id", ""
                    )
                    worksheet[f"{output_columns['data_type']}{row}"] = match_result.get(
                        "data_type", ""
                    )
                    worksheet[f"{output_columns['length_total']}{row}"] = match_result.get(
                        "length_total", ""
                    )
                    worksheet[f"{output_columns['length_dec']}{row}"] = match_result.get(
                        "length_dec", ""
                    )
                    worksheet[f"{output_columns['match']}{row}"] = match_result.get(
                        "match", ""
                    )
                    worksheet[f"{output_columns['notes']}{row}"] = match_result.get(
                        "notes", ""
                    )
                    worksheet[f"{output_columns['sample_value']}{row}"] = match_result.get(
                        "sample_value", ""
                    )
                    worksheet[f"{output_columns['verify']}{row}"] = match_result.get(
                        "verify", ""
                    )

                    processed_count += 1

                except Exception as e:
                        continue

    # ========== Field Matching Logic ==========

    def _match_fields(
        self,
        input_fields: List,
        context: Optional[List[Dict[str, Any]]] = None,
        excel_filename: str = None,
    ) -> List[Dict[str, Any]]:
        """Field matching using AI services with HANA context"""
        if not input_fields:
            return []

        if context is None:
            return []

        return self._match_fields_with_context(input_fields, context, excel_filename)

    def _match_fields_with_context(
        self,
        input_fields: List,
        context: List[Dict[str, Any]],
        excel_filename: str = None,
    ) -> List[Dict[str, Any]]:
        all_fields_context = self._extract_fields_from_context(context)

        resulsts_prompt = self.ai_service.get_rag_matching_prompt(
            input_fields, all_fields_context
        )

        results_function_schema = self.ai_service.get_field_matching_schema()

        results_response = self.ai_service.call_with_function(
            resulsts_prompt, results_function_schema
        )

        # Check if results_response is empty or None
        if not results_response:
            error_msg = _("LLM returned empty response for field matching")
            logger.error(
                error_msg,
                logger.get_excel_log_filename(excel_filename)
                if excel_filename
                else None,
            )
            raise RuntimeError(error_msg)

        return self._parse_llm_response(results_response, input_fields, excel_filename)

    def _extract_fields_from_context(
        self, context: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """Extract all fields from the matched views as context."""
        all_fields_context = []

        for field_detail in context:
            all_fields_context.append(
                {
                    "view_name": field_detail.get("view_name", ""),
                    "view_desc": field_detail.get("view_desc", ""),
                    "field_name": str(field_detail.get("field_name", "")),
                    "field_desc": str(field_detail.get("field_desc", "")),
                    "is_key": bool(field_detail.get("is_key", False)),
                    "data_type": str(field_detail.get("data_type", "")),
                    "field_id": str(field_detail.get("field_id", "")),
                    "length_total": str(field_detail.get("length_total", "")),
                    "length_dec": str(field_detail.get("length_dec", "")),
                }
            )

        return all_fields_context

    def _parse_llm_response(
        self,
        function_response: Dict[str, Any],
        input_fields: List,
        excel_filename: str = None,
    ) -> List[Dict[str, Any]]:
        results = []
        matches = function_response.get("review", [])

        # Create mapping by row_index for exact matching
        match_map = {
            m.get("row_index"): m for m in matches if m.get("row_index") is not None
        }

        # Log matching info for debugging
        logger.debug(
            f"LLM returned {len(matches)} matches for {len(input_fields)} input fields",
            logger.get_excel_log_filename(excel_filename) if excel_filename else None,
        )

        for input_field in input_fields:
            # Handle both InterfaceField objects and dictionaries
            if hasattr(input_field, "row_index"):
                # InterfaceField object
                row_index = input_field.row_index
            else:
                # Dictionary
                row_index = input_field.get("row_index")

            match_result = match_map.get(row_index, {})

            # If no exact row match found, log warning
            if not match_result:
                logger.warning(
                    f"⚠️ No LLM match found for row {row_index}, using empty values",
                    logger.get_excel_log_filename(excel_filename)
                    if excel_filename
                    else None,
                )

            key_flag_raw = match_result.get("key_flag", "")
            if isinstance(key_flag_raw, bool):
                key_flag = "○" if key_flag_raw else ""
            elif isinstance(key_flag_raw, str):
                key_flag = (
                    "○" if key_flag_raw.lower() in ["true", "y", "yes", "x"] else ""
                )
            else:
                key_flag = ""

            # Clean field_id to remove view name prefix (technical field name)
            raw_field_id = match_result.get("field_id", "")
            clean_field_id = ""

            if raw_field_id:
                # Remove view name prefix (e.g., "I_TIMESHEETRECORD.RECEIVERCOSTCENTER" -> "RECEIVERCOSTCENTER")
                if "." in raw_field_id:
                    clean_field_id = raw_field_id.split(".")[-1]
                else:
                    clean_field_id = raw_field_id

            # Parse notes to extract percentage and description
            notes_text = match_result.get("notes", "")
            match_percentage, notes_description = self._parse_notes(notes_text)

            # field_name should be the field description
            field_name = match_result.get("field_desc", "")

            results.append(
                {
                    "table_id": match_result.get("table_id", ""),
                    "field_id": clean_field_id,  # Technical field name
                    "field_name": field_name,  # Field description
                    "key_flag": key_flag,
                    "obligatory": match_result.get("obligatory", ""),
                    "data_type": match_result.get("data_type", ""),
                    "length_total": match_result.get("length_total", ""),
                    "length_dec": match_result.get("length_dec", ""),
                    "field_desc": match_result.get("field_desc", ""),
                    "sample_value": match_result.get("sample_value", ""),
                    "match": match_result.get("match"),
                    "notes": match_result.get("notes", ""),
                }
            )

        # Call OData service to verify results
        results = odata_verify(results)

        return results

    def _parse_notes(self, notes_text: str) -> Tuple[str, str]:
        """Parse notes text to extract percentage and description.

        Expected format: "XX% - Description" or "XX%: Description" or "Description (XX%)"
        Returns: (percentage, description)
        """
        if not notes_text:
            return "", ""

        import re

        # Pattern 1: "XX% - Description" or "XX%: Description"
        pattern1 = r"^(\d+%)\s*[-:]\s*(.+)$"
        match1 = re.match(pattern1, notes_text.strip())
        if match1:
            return match1.group(1), match1.group(2).strip()

        # Pattern 2: "Description (XX%)"
        pattern2 = r"^(.+)\s*\((\d+%)\)\s*$"
        match2 = re.match(pattern2, notes_text.strip())
        if match2:
            return match2.group(2), match2.group(1).strip()

        # Pattern 3: Just percentage "XX%"
        pattern3 = r"^(\d+%)\s*$"
        match3 = re.match(pattern3, notes_text.strip())
        if match3:
            return match3.group(1), ""

        # Pattern 4: Contains percentage anywhere
        pattern4 = r"(\d+%)"
        match4 = re.search(pattern4, notes_text)
        if match4:
            percentage = match4.group(1)
            # Remove percentage from description
            description = re.sub(r"\s*\d+%\s*[-:()]*\s*", " ", notes_text).strip()
            return percentage, description

        # No percentage found, return everything as description
        return "", notes_text.strip()

    def _archive_processed_file(self, source_file_path: Path) -> bool:
        """Move successfully processed file to excel_archive folder.

        Args:
            source_file_path: Path to the source file to archive

        Returns:
            bool: True if archiving was successful, False otherwise
        """
        try:
            from core.consts import Directories

            # Create archive directory if it doesn't exist
            archive_dir = self.data_dir / Directories.EXCEL_ARCHIVE
            archive_dir.mkdir(exist_ok=True)

            # Generate archived filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            archive_filename = f"{timestamp}_{source_file_path.name}"
            archive_path = archive_dir / archive_filename

            # Move the file to archive
            shutil.move(str(source_file_path), str(archive_path))

            return True

        except Exception as e:
            logger.error(
                f"❌ Failed to archive file {source_file_path.name}: {e}",
                logger.get_excel_log_filename(source_file_path.name),
            )
            return False
               