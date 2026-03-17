import json
import os
from typing import List, Dict, Any, Tuple

import pandas as pd
from dotenv import load_dotenv
from hana_ml import ConnectionContext
from hdbcli.dbapi import Error as HanaDbError

from utils.i18n import _, get_current_language
from utils.sap_logger import logger
from collections import defaultdict


load_dotenv()


class HANADBClient:
    def __init__(self):
        self.scenario_table = "PWC_HAND_AI2REPORT_DEV_BUSINESSSCENARIOS"
        self.cds_view_table = "PWC_HAND_AI2REPORT_DEV_CDSVIEWS"
        self.view_fields_table = "PWC_HAND_AI2REPORT_DEV_VIEWFIELDS"
        self.cust_fields_table = "PWC_HAND_AI2REPORT_DEV_CUSTFIELDS"
        self.TerminologyMapping = "PWC_HAND_AI2REPORT_DEV_TERMINOLOGYMAPPING"
        
        self.db_addr = os.getenv("HANA_ADDRESS")
        self.db_user = os.getenv("HANA_USER")
        self.db_pwd = os.getenv("HANA_PASSWORD")
        self._db_schema = os.getenv("HANA_SCHEMA")  
        self._db_schema_cust = os.getenv("HANA_SCHEMA_CUST")

        self.hana_client: ConnectionContext = None

    def connect(self) -> None:
        if self.hana_client:
            return

        try:
            self.hana_client = ConnectionContext(
                address=self.db_addr,
                port="443",  # 443 is usual
                user=self.db_user,
                password=self.db_pwd,
                encrypt=True,
            )
        except HanaDbError as e:
            logger.error(_("HANA Cloud connection failed: {}").format(e))
            raise

    def close(self) -> None:
        """Close database connection."""
        if self.hana_client:
            self.hana_client.close()
            logger.info(_("Database connection closed."))

    def _format_in_clause(self, items: List[str]) -> str:
        if not items:
            return "('')"
        formatted_items = [f"'{str(item).replace("'", "''")}'" for item in items]
        return f"({', '.join(formatted_items)})"

    def run_vector_search(
        self,
        query: str,
        metric="COSINE_SIMILARITY",
        k=3,
        log_filename: str = None,
    ) -> pd.DataFrame:
        if not self.hana_client:
            error_msg = _("HANA Cloud not connected.")
            logger.error(error_msg, log_filename)
            raise ConnectionError(error_msg)

        sort = "ASC" if metric == "L2DISTANCE" else "DESC"

        sql = """SELECT TOP {k}"ID","SCENARIO","DESCRIPTION","VIEWCATEGORY"
        FROM "{schema}"."{table}"
        ORDER BY {metric}(VECTOR_EMBEDDING('{query}', 'QUERY', 'SAP_NEB.20240715'),"EMBEDDINGS") {sort}""".format(
            k=k,
            metric=metric,
            query=query,
            sort=sort,
            table=self.scenario_table,
            schema=self._db_schema,
        )
        try:
            hdf = self.hana_client.sql(sql)
            df_context = hdf.head(k).collect()
            return df_context
        except HanaDbError as e:
            logger.error(_("SQL execution failed: {}").format(e), log_filename)
            return pd.DataFrame()

    def get_views(self, category: str, log_filename: str = None) -> pd.DataFrame:
        if not self.hana_client:
            error_msg = _("Database not connected.")
            logger.error(error_msg, log_filename)
            raise ConnectionError(error_msg)

        categories = [cat for cat in category.split("/") if cat]
        if not categories:
            logger.warning(_("No valid categories provided"), log_filename)
            return pd.DataFrame()

        category_sql = self._format_in_clause(categories)

        sql = """
             SELECT "VIEWNAME", "VIEWDESC"
            FROM "{schema}"."{table}"
            WHERE "VIEWCATEGORY" IN {categories} 
            AND "ISACTIVE" = 'true'
        """.format(
            schema=self._db_schema, table=self.cds_view_table, categories=category_sql
        )
        try:
            result = self.hana_client.sql(sql).collect()
            return result
        except HanaDbError as e:
            logger.error(
                _("SQL error occurred while getting views: {}").format(e), log_filename
            )
            return pd.DataFrame()

    def get_fields(
        self, cds_views: List[str], log_filename: str = None
    ) -> Dict[str, List[Dict[str, Any]]]:
        cds_views_sql = self._format_in_clause(cds_views)

        sql = """
            SELECT "TABLENAME","TABLEDESC","CONTENT"
            FROM "{schema}"."{table}"
            WHERE "TABLENAME" IN {cds_views}
            AND "LANGU" = 'ja'
        """.format(
            schema=self._db_schema,
            table=self.view_fields_table,
            cds_views=cds_views_sql,
        )
        try:
            fields_df = self.hana_client.sql(sql).collect()
            # 初始化结果字典
            results = {view: [] for view in cds_views}
            if not fields_df.empty:
                for _, row in fields_df.iterrows():
                    view_name = row["TABLENAME"]
                    content_str = row["CONTENT"]

                    if not content_str or not isinstance(content_str, str):
                        continue

                    try:
                        # 2. 解析字段
                        content_str_re = self.parse_fields(content_str)
                        # 直接找到[[的位置
                        start_idx = content_str_re.find('[[')
                        end_idx = content_str_re.rfind(']]')

                        if start_idx != -1 and end_idx != -1:
                             # 提取完整的内容（包含[[和]]）
                            full_content = content_str_re[start_idx:end_idx+2]
                            
                        parsed_fields = json.loads(full_content)
                        
                        # 3. 将解析后的列表转换为结构化的字典列表
                        for field_data in parsed_fields:
                            if not isinstance(field_data, list) or len(field_data) < 7:
                                continue  # Skip malformed entries

                            field_dict = {
                                "field_name": field_data[0],
                                "is_key": field_data[1],
                                "field_desc": field_data[2],
                                "data_element": field_data[3],  # Can be added if needed
                                "data_type": field_data[4],
                                "length_total": field_data[5],
                                "length_dec": field_data[6],
                            }
                            results[view_name].append(field_dict)
                    except json.JSONDecodeError as e:
                        logger.warning(
                            _(
                                "⚠️ Warning: Could not parse CONTENT JSON for view {} error: {}"
                            ).format(view_name, e),
                            log_filename,
                        )
                        continue
                    except Exception as e:
                        logger.error(_("Error parse fields:{}").format(e), log_filename)
            return results
        except HanaDbError as e:
            logger.error(_("Error: {}").format(e), log_filename)
            return {}

    def _build_custom_field_result(self, row) -> Dict[str, Any]:
        """将 CUSTFIELDS 表的一行转换为统一的结果字典"""
        return {
            "table_id":     row["TARGETTABLE"],
            "field_id":     row["TARGETFIELD"],
            "field_name":   row["TARGETDESC"],
            "data_type":    row["TARGETTYPE"],
            "length_total": str(row["TARGETLENGTH"]) if row["TARGETLENGTH"] is not None else "",
            "length_dec":   str(row["TARGETDECIMALS"]) if row["TARGETDECIMALS"] is not None else "",
            "key_flag":     row["KEYFLAG"],
            "obligatory":   row["OBLIGATORY"],
            "sample_value": row["ALLOWEDVALUES"],
            "notes":        row["NOTES"],
            "color":        row["COLOR"] if row["COLOR"] else "",
        }

    def get_custom_fields_exact(
        self,
        source_table: str,
        source_field: str,
        log_filename: str = None,
    ) -> Tuple[Dict[str, Any], bool]:
        """
        第一步：精确匹配，根据 SourceTable + SourceField 查询 CUSTFIELDS 表。

        Args:
            source_table: 输入字段的 table_id
            source_field: 输入字段的 field_id
            log_filename: 日志文件名

        Returns:
            (匹配结果字典, is_multiple: bool)
            - 唯一匹配时：(result_dict, False)
            - 无匹配时：  ({}, False)
            - 多条匹配时：({}, True)  ← 调用方应据此缩小向量检索范围
        """

        clean_table = source_table.replace("'", "''") if source_table else ""
        clean_field = source_field.replace("'", "''") if source_field else ""

        # 空值用 IS NULL 匹配，非空值用 = 匹配
        def _null_or_eq(col: str, val: str) -> str:
            if not val or val.strip() in ("", "-"):
                return f'("{col}" IS NULL OR "{col}" = \'\')'
            return f'"{col}" = \'{val}\''

        table_cond = _null_or_eq("SOURCETABLE", clean_table)
        field_cond  = _null_or_eq("SOURCEFIELD",  clean_field)

        sql = """
            SELECT TOP 2
                "TARGETTABLE", "TARGETFIELD", "TARGETDESC",
                "TARGETTYPE", "TARGETLENGTH", "TARGETDECIMALS",
                "KEYFLAG", "OBLIGATORY", "ALLOWEDVALUES", "NOTES", "COLOR"
            FROM "{schema}"."{table}"
            WHERE {table_cond}
              AND {field_cond}
              AND "ISACTIVE" = 0
        """.format(
            schema=self._db_schema_cust,
            table=self.cust_fields_table,
            table_cond=table_cond,
            field_cond=field_cond,
        )

        try:
            result_df = self.hana_client.sql(sql).collect()
            if result_df.empty:
                return {}, False
            if len(result_df) > 1:
                return {}, True   # multiple rows found → caller should scope vector search
            return self._build_custom_field_result(result_df.iloc[0]), False
        except HanaDbError as e:
            logger.error(
                _("Custom field exact match failed: {}").format(e),
                log_filename,
            )
            return {}, False

    def get_custom_fields(
        self,
        field_query: str,
        metric="COSINE_SIMILARITY",
        log_filename: str = None,
        source_table: str = None,
        source_field: str = None,
    ) -> Dict[str, Any]:
        """
        第二步：向量匹配，将拼接后的查询文本与 CUSTFIELDS 表的 embeddings 进行相似度检索。
        embeddings 由 IFName + SourceTable + SourceField + SourceDesc 拼接向量化得到。

        Args:
            field_query: 查询文本（IFName + SourceTable + SourceField + SourceDesc 拼接）
            metric: 相似度算法
            log_filename: 日志文件名
            source_table: 若精确匹配返回多条，传入以缩小检索范围
            source_field: 若精确匹配返回多条，传入以缩小检索范围

        Returns:
            匹配结果字典，未匹配返回空字典
        """
        sort = "ASC" if metric == "L2DISTANCE" else "DESC"
        comparison_op = ">" if sort.upper() == "DESC" else "<"
        threshold = float(os.getenv("CUSTOM_FIELD_THRESHOLD", 0.75))

        clean_query = field_query.replace("'", "''").replace('"', '""')

        # 当精确匹配命中多条时，追加 SOURCETABLE / SOURCEFIELD 过滤条件（兼容 NULL）
        def _scope_cond(col: str, val: str) -> str:
            if not val or val.strip() in ("", "-"):
                return f'("{col}" IS NULL OR "{col}" = \'\')'
            return f'"{col}" = \'{val.replace(chr(39), chr(39)*2)}\''

        scope_filter = ""
        if source_table is not None:
            scope_filter += f' AND {_scope_cond("SOURCETABLE", source_table)}'
        if source_field is not None:
            scope_filter += f' AND {_scope_cond("SOURCEFIELD", source_field)}'

        sql = """
            SELECT TOP 1
                "TARGETTABLE", "TARGETFIELD", "TARGETDESC",
                "TARGETTYPE", "TARGETLENGTH", "TARGETDECIMALS",
                "KEYFLAG", "OBLIGATORY", "ALLOWEDVALUES", "NOTES", "COLOR",
                SIMILARITY_SCORE
            FROM (
                SELECT
                    "TARGETTABLE", "TARGETFIELD", "TARGETDESC",
                    "TARGETTYPE", "TARGETLENGTH", "TARGETDECIMALS",
                    "KEYFLAG", "OBLIGATORY", "ALLOWEDVALUES", "NOTES", "COLOR",
                    {metric}(VECTOR_EMBEDDING('{query}', 'QUERY', 'SAP_NEB.20240715'), "EMBEDDINGS") AS SIMILARITY_SCORE
                FROM "{schema}"."{table}"
                WHERE "ISACTIVE" = 0{scope_filter}
            ) AS T
            WHERE SIMILARITY_SCORE {op} {threshold}
            ORDER BY SIMILARITY_SCORE {sort}
        """.format(
            metric=metric,
            query=clean_query,
            sort=sort,
            op=comparison_op,
            threshold=threshold,
            schema=self._db_schema_cust,
            table=self.cust_fields_table,
            scope_filter=scope_filter,
        )

        try:
            result_df = self.hana_client.sql(sql).collect()
            if result_df.empty:
                return {}
            return self._build_custom_field_result(result_df.iloc[0])
        except HanaDbError as e:
            logger.error(
                _("Custom field vector search failed: {}").format(e),
                log_filename,
            )
            return {}


    def upload_custfields_from_excel(
        self,
        excel_path: str,
        sheet_name: str = None,
        log_filename: str = None,
    ) -> Dict[str, int]:
        """
        Excel の対応表を読み込み、CUSTFIELDS テーブルへ upsert する。
        content = IFNAME + SOURCETABLE + SOURCEFIELD + SOURCEDESC を結合してセット。
        embeddings は DB の trigger が content を元に自動生成する。
        """
        import openpyxl

        if not self.hana_client:
            raise ConnectionError(_("Database not connected."))

        wb = openpyxl.load_workbook(excel_path)

        if sheet_name:
            ws = wb[sheet_name]
        else:
            preferred = [s for s in wb.sheetnames if "正本" in s]
            ws = wb[preferred[0]] if preferred else wb.active

        logger.info(
            _("Uploading from sheet '{}' in '{}'").format(ws.title, excel_path),
            log_filename,
        )

        def _cell_hex_color(cell) -> str:
            try:
                fill = cell.fill
                if fill and fill.fill_type not in (None, "none"):
                    fg = fill.fgColor
                    if fg.type == "rgb":
                        argb = fg.rgb
                        if argb and argb != "00000000":
                            return "#" + argb[-6:]
            except Exception:
                pass
            return ""

        def _val(v):
            if v is None:
                return "NULL"
            return "'" + str(v).replace("'", "''") + "'"

        def _eq_or_null(col, val):
            if not val:
                return f'("{col}" IS NULL OR "{col}" = \'\')'
            return f'"{col}" = \'{val.replace(chr(39), chr(39)*2)}\''

        rows_data = []
        for row in ws.iter_rows(min_row=2):
            if_name      = row[1].value
            source_desc  = row[2].value
            source_table = row[3].value
            source_field = row[4].value
            target_desc  = row[5].value
            target_table = row[6].value
            target_field = row[7].value
            notes        = row[8].value if len(row) > 8 else None
            color        = _cell_hex_color(row[5])

            if not source_field or str(source_field).strip() in ("", "e"):
                continue

            st = str(source_table or "").strip() if source_table else None
            sf = str(source_field or "").strip() if source_field else None

            content = " ".join(filter(None, [
                str(if_name or "").strip(),
                st or "",
                sf or "",
                str(source_desc or "").strip(),
            ]))

            rows_data.append({
                "IFNAME":      str(if_name or "").strip(),
                "SOURCEDESC":  str(source_desc or "").strip(),
                "SOURCETABLE": st,
                "SOURCEFIELD": sf,
                "TARGETDESC":  str(target_desc or "").strip(),
                "TARGETTABLE": str(target_table or "").strip(),
                "TARGETFIELD": str(target_field or "").strip(),
                "NOTES":       str(notes or "").strip(),
                "COLOR":       color,
                "CONTENT":     content,
            })

        logger.info(_("Found {} rows to process").format(len(rows_data)), log_filename)

        upload_mode = os.getenv("UPLOAD_MODE", "upsert").strip().lower()
        logger.info(_("Upload mode: {}").format(upload_mode), log_filename)

        stats = {"inserted": 0, "updated": 0, "skipped": 0, "errors": 0}

        # ── overwrite モード：既存レコードを全削除してから INSERT ────────────
        if upload_mode == "overwrite":
            try:
                delete_sql = """
                    DELETE FROM "{schema}"."{table}" WHERE "ISACTIVE" = 0
                """.format(
                    schema=self._db_schema_cust,
                    table=self.cust_fields_table,
                )
                cursor = self.hana_client.connection.cursor()
                cursor.execute(delete_sql)
                cursor.close()
                logger.info(_("Overwrite mode: existing records deleted."), log_filename)
            except HanaDbError as e:
                logger.error(_("Failed to delete existing records: {}").format(e), log_filename)
                raise

        for idx, row in enumerate(rows_data):
            try:
                if upload_mode == "overwrite":
                    # overwrite モードは常に INSERT
                    insert_sql = """
                        INSERT INTO "{schema}"."{table}"
                            ("ID", "IFNAME", "SOURCEDESC", "SOURCETABLE", "SOURCEFIELD",
                             "TARGETDESC", "TARGETTABLE", "TARGETFIELD",
                             "NOTES", "COLOR", "ISACTIVE", "CONTENT")
                        VALUES (
                            SYSUUID,
                            {ifname}, {sdesc}, {stable}, {sfield},
                            {tdesc}, {ttable}, {tfield},
                            {notes}, {color}, 0, {content}
                        )
                    """.format(
                        schema=self._db_schema_cust,
                        table=self.cust_fields_table,
                        ifname=_val(row["IFNAME"]),
                        sdesc=_val(row["SOURCEDESC"]),
                        stable=_val(row["SOURCETABLE"]),
                        sfield=_val(row["SOURCEFIELD"]),
                        tdesc=_val(row["TARGETDESC"]),
                        ttable=_val(row["TARGETTABLE"]),
                        tfield=_val(row["TARGETFIELD"]),
                        notes=_val(row["NOTES"]),
                        color=_val(row["COLOR"]),
                        content=_val(row["CONTENT"]),
                    )
                    cursor = self.hana_client.connection.cursor()
                    cursor.execute(insert_sql)
                    cursor.close()
                    stats["inserted"] += 1

                else:
                    # upsert モード：既存確認 → UPDATE or INSERT
                    check_sql = """
                        SELECT "ID" FROM "{schema}"."{table}"
                        WHERE {st} AND {sf} AND "TARGETFIELD" = '{tf}'
                        AND "ISACTIVE" = 0
                    """.format(
                        schema=self._db_schema_cust,
                        table=self.cust_fields_table,
                        st=_eq_or_null("SOURCETABLE", row["SOURCETABLE"]),
                        sf=_eq_or_null("SOURCEFIELD", row["SOURCEFIELD"]),
                        tf=(row["TARGETFIELD"] or "").replace("'", "''"),
                    )
                    existing = self.hana_client.sql(check_sql).collect()

                    if not existing.empty:
                        record_id = existing.iloc[0]["ID"]
                        # CAP managed entity 的 UPDATE 会触发钩子导致新增记录
                        # 改为先 DELETE 再 INSERT 以避免此问题
                        delete_sql = """
                            DELETE FROM "{schema}"."{table}" WHERE "ID" = '{rid}'
                        """.format(
                            schema=self._db_schema_cust,
                            table=self.cust_fields_table,
                            rid=record_id,
                        )
                        insert_sql = """
                            INSERT INTO "{schema}"."{table}"
                                ("ID", "IFNAME", "SOURCEDESC", "SOURCETABLE", "SOURCEFIELD",
                                 "TARGETDESC", "TARGETTABLE", "TARGETFIELD",
                                 "NOTES", "COLOR", "ISACTIVE", "CONTENT")
                            VALUES (
                                '{rid}',
                                {ifname}, {sdesc}, {stable}, {sfield},
                                {tdesc}, {ttable}, {tfield},
                                {notes}, {color}, 0, {content}
                            )
                        """.format(
                            schema=self._db_schema_cust,
                            table=self.cust_fields_table,
                            rid=record_id,
                            ifname=_val(row["IFNAME"]),
                            sdesc=_val(row["SOURCEDESC"]),
                            stable=_val(row["SOURCETABLE"]),
                            sfield=_val(row["SOURCEFIELD"]),
                            tdesc=_val(row["TARGETDESC"]),
                            ttable=_val(row["TARGETTABLE"]),
                            tfield=_val(row["TARGETFIELD"]),
                            notes=_val(row["NOTES"]),
                            color=_val(row["COLOR"]),
                            content=_val(row["CONTENT"]),
                        )
                        cursor = self.hana_client.connection.cursor()
                        cursor.execute(delete_sql)
                        cursor.execute(insert_sql)
                        cursor.close()
                        stats["updated"] += 1
                    else:
                        insert_sql = """
                            INSERT INTO "{schema}"."{table}"
                                ("ID", "IFNAME", "SOURCEDESC", "SOURCETABLE", "SOURCEFIELD",
                                 "TARGETDESC", "TARGETTABLE", "TARGETFIELD",
                                 "NOTES", "COLOR", "ISACTIVE", "CONTENT")
                            VALUES (
                                SYSUUID,
                                {ifname}, {sdesc}, {stable}, {sfield},
                                {tdesc}, {ttable}, {tfield},
                                {notes}, {color}, 0, {content}
                            )
                        """.format(
                            schema=self._db_schema_cust,
                            table=self.cust_fields_table,
                            ifname=_val(row["IFNAME"]),
                            sdesc=_val(row["SOURCEDESC"]),
                            stable=_val(row["SOURCETABLE"]),
                            sfield=_val(row["SOURCEFIELD"]),
                            tdesc=_val(row["TARGETDESC"]),
                            ttable=_val(row["TARGETTABLE"]),
                            tfield=_val(row["TARGETFIELD"]),
                            notes=_val(row["NOTES"]),
                            color=_val(row["COLOR"]),
                            content=_val(row["CONTENT"]),
                        )
                        cursor = self.hana_client.connection.cursor()
                        cursor.execute(insert_sql)
                        cursor.close()
                        stats["inserted"] += 1

            except HanaDbError as e:
                logger.error(
                    _("Row {}: DB error - {}").format(idx + 2, e), log_filename
                )
                stats["errors"] += 1

        logger.info(
            _("Upload complete: inserted={inserted}, updated={updated}, "
              "skipped={skipped}, errors={errors}").format(**stats),
            log_filename,
        )
        return stats


    def get_terms(self,log_filename: str = None) -> pd.DataFrame:
        if not self.hana_client:
            error_msg = _("Database not connected.")
            logger.error(error_msg, log_filename)
            raise ConnectionError(error_msg)

        sql = """
             SELECT "SOURCETERM", 
                    "SOURCETERMALIAS", 
                    "SOURCECONTEXT",
                    "TARGETTERM", 
                    "TARGETTERMALIAS",
                    "SAPMODULE",
                    "SAPTRANSACTION",
                    "SAPOBJECTTYPE",
                    "SAPTECHNICALNAME",
                    "CATEGORY",
                    "DOMAINAREA",
                    "PRIORITY",
                    "CONFIDENCE"
               FROM "{schema}"."{table}"
              WHERE "STATUS" = 'ACTIVE'
        """.format(
            schema=self._db_schema, table=self.TerminologyMapping
        )
        try:
            result = self.hana_client.sql(sql).collect()
            return result
        except HanaDbError as e:
            logger.error(
                _("SQL error occurred while getting views: {}").format(e), log_filename
            )
            return pd.DataFrame()
            
    @staticmethod
    def parse_fields(content_str: str) -> str:
        result_chars = []
        in_string = False
        i = 0
        s_len = len(content_str)

        while i < s_len:
            char = content_str[i]

            if char == '"':
                if not in_string:
                    # 字符串开头
                    in_string = True
                    result_chars.append(char)
                else:
                    # 字符串内部遇到了一个引号。
                    # 查找下一个非空白字符
                    next_char_index = i + 1
                    while (
                        next_char_index < s_len
                        and content_str[next_char_index].isspace()
                    ):
                        next_char_index += 1

                    # 如果字符串后面就是逗号、方括号或字符串结尾，
                    # 那么这个引号是合法的结束符。
                    if next_char_index == s_len or content_str[next_char_index] in [
                        ",",
                        "]",
                        "}",
                    ]:
                        in_string = False
                        result_chars.append(char)
                    else:
                        # 否则，这绝对是一个需要转义的内部引号。
                        result_chars.append('\\"')
            else:
                # 对于所有其他字符，直接添加
                result_chars.append(char)

            i += 1

        repaired_string = "".join(result_chars)

        return repaired_string


if __name__ == "__main__":
    query_text = "purchase"
    try:
        db = HANADBClient()
        db.connect()
        
        logger.info(
            _("--- Step 1: Executing vector search for query '{}' ---").format(
                query_text
            )
        )
        vector_search_results = db.run_vector_search(query=query_text, k=1)

        logger.info(_("\nVector search results:"))
        logger.info(str(vector_search_results))
        logger.info("-" * 80)

        if not vector_search_results.empty:
            category_string = vector_search_results.iloc[0]["VIEWCATEGORY"]
            logger.info(
                _(
                    "\n--- Step 2: Getting all related CDS views using category string '{}' ---"
                ).format(category_string)
            )

            views_in_category = db.get_views(category=category_string)

            if not views_in_category.empty:
                logger.info(
                    _("\nFound {} CDS views:").format(len(views_in_category))
                )
                logger.info("-" * 80)

        else:
            logger.info(_("\nVector search returned no results."))
    except Exception as e:
        logger.error(_("\nProgram execution failed: {}").format(e))
